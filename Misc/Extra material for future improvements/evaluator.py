"""
build_baseline_excel.py
=======================
Reads grid_manifest_op1.csv and grid_manifest_op2.csv,
parses the corresponding LLM .txt output files,
runs the FULL MILP optimizer (both PV-only and PV+Battery) for every row,
fetches real irradiance from Open-Meteo API (cached by lat/lon so each unique
location is only fetched once across all 300+ rows),
and writes TWO Excel files:

  baseline_comparison.xlsx   — full detail (all cols)
  summary_comparison.xlsx    — CAPEX + Annual Savings for all 4 systems

Run:
    python build_baseline_excel.py

Adjust the PATH constants at the top.
"""

import hashlib
import math
import os
import re
import time
import warnings
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.optimize import milp, LinearConstraint, Bounds
from scipy.sparse import csc_matrix

warnings.filterwarnings("ignore")


# ══════════════════════════════════════════════════════════════════════════════
# PATH CONFIG
# ══════════════════════════════════════════════════════════════════════════════

MANIFEST_OP1  = r"C:\Users\shubh\Desktop\Hard disk\College(PG)\Academics at UCSD\Y1Q2\ECE 285 - Spec Topic - Signal & Image Robotics - Smartgrids\Project\Step 4 Track B\Evaluation against baseline\output_1\grid_manifest_op1.csv"
MANIFEST_OP2  = r"C:\Users\shubh\Desktop\Hard disk\College(PG)\Academics at UCSD\Y1Q2\ECE 285 - Spec Topic - Signal & Image Robotics - Smartgrids\Project\Step 4 Track B\Evaluation against baseline\output_2\grid_manifest_op2.csv"
OUTPUT_DIR_1  = r"C:\Users\shubh\Desktop\Hard disk\College(PG)\Academics at UCSD\Y1Q2\ECE 285 - Spec Topic - Signal & Image Robotics - Smartgrids\Project\Step 4 Track B\Evaluation against baseline\output_1"
OUTPUT_DIR_2  = r"C:\Users\shubh\Desktop\Hard disk\College(PG)\Academics at UCSD\Y1Q2\ECE 285 - Spec Topic - Signal & Image Robotics - Smartgrids\Project\Step 4 Track B\Evaluation against baseline\output_2"
TOU_DR_PATH   = r"C:\Users\shubh\Desktop\Hard disk\College(PG)\Academics at UCSD\Y1Q2\ECE 285 - Spec Topic - Signal & Image Robotics - Smartgrids\Project\Step 1 Preprocess data\Tariff Data\TOU-DR\tou_dr_daily_2021_2025.csv"
TOU_DR1_PATH  = r"C:\Users\shubh\Desktop\Hard disk\College(PG)\Academics at UCSD\Y1Q2\ECE 285 - Spec Topic - Signal & Image Robotics - Smartgrids\Project\Step 1 Preprocess data\Tariff Data\TOU-DR1\tou_dr1_daily_2021_2025.csv"
TOU_DR2_PATH  = r"C:\Users\shubh\Desktop\Hard disk\College(PG)\Academics at UCSD\Y1Q2\ECE 285 - Spec Topic - Signal & Image Robotics - Smartgrids\Project\Step 1 Preprocess data\Tariff Data\TOU-DR2\tou_dr2_daily_2021_2025.csv"
EIA_LOAD_PATH = r"C:\Users\shubh\Desktop\Hard disk\College(PG)\Academics at UCSD\Y1Q2\ECE 285 - Spec Topic - Signal & Image Robotics - Smartgrids\Project\Step 1 Preprocess data\San_Diego_Load_EIA_Fixed.csv"
OUTPUT_XLSX   = r"C:\Users\shubh\Desktop\Hard disk\College(PG)\Academics at UCSD\Y1Q2\ECE 285 - Spec Topic - Signal & Image Robotics - Smartgrids\Project\Step 4 Track B\Evaluation against baseline\baseline_comparison.xlsx"
SUMMARY_XLSX  = r"C:\Users\shubh\Desktop\Hard disk\College(PG)\Academics at UCSD\Y1Q2\ECE 285 - Spec Topic - Signal & Image Robotics - Smartgrids\Project\Step 4 Track B\Evaluation against baseline\summary_comparison.xlsx"


# ══════════════════════════════════════════════════════════════════════════════
# CATALOGS AND CONSTANTS
# ══════════════════════════════════════════════════════════════════════════════

SOLAR_PANEL_CATALOG = [
    {"manufacturer": "REC Group",      "model": "Alpha Pure",
     "efficiency_percent": 21.9, "cost_per_wp_usd": 0.83, "temp_coeff_pct_per_C": -0.26,
     "voc_v": 49.1,  "isc_a": 10.41, "vmp_v": 41.8,  "imp_a": 9.69,
     "panel_power_w": 405, "panel_area_m2": 1.85,
     "cells_in_series": 66, "cells_in_parallel": 2, "degradation_rate": 0.0025},
    {"manufacturer": "JA Solar",       "model": "DeepBlue 3.0",
     "efficiency_percent": 20.2, "cost_per_wp_usd": 0.45, "temp_coeff_pct_per_C": -0.35,
     "voc_v": 36.98, "isc_a": 13.7,  "vmp_v": 30.84, "imp_a": 12.81,
     "panel_power_w": 395, "panel_area_m2": 1.95,
     "cells_in_series": 54, "cells_in_parallel": 2, "degradation_rate": 0.0061},
    {"manufacturer": "Trina Solar",    "model": "Vertex S",
     "efficiency_percent": 20.8, "cost_per_wp_usd": 0.32, "temp_coeff_pct_per_C": -0.34,
     "voc_v": 41.2,  "isc_a": 12.28, "vmp_v": 34.2,  "imp_a": 11.7,
     "panel_power_w": 400, "panel_area_m2": 1.92,
     "cells_in_series": 60, "cells_in_parallel": 2, "degradation_rate": 0.0055},
    {"manufacturer": "Canadian Solar", "model": "TOPHiKu7",
     "efficiency_percent": 23.2, "cost_per_wp_usd": 0.16, "temp_coeff_pct_per_C": -0.29,
     "voc_v": 48.7,  "isc_a": 18.69, "vmp_v": 40.8,  "imp_a": 17.67,
     "panel_power_w": 720, "panel_area_m2": 3.11,
     "cells_in_series": 66, "cells_in_parallel": 2, "degradation_rate": 0.004},
    {"manufacturer": "Silfab Solar",   "model": "Prime",
     "efficiency_percent": 22.6, "cost_per_wp_usd": 0.53, "temp_coeff_pct_per_C": -0.36,
     "voc_v": 38.97, "isc_a": 14.22, "vmp_v": 33.41, "imp_a": 13.17,
     "panel_power_w": 440, "panel_area_m2": 1.94,
     "cells_in_series": 54, "cells_in_parallel": 2, "degradation_rate": 0.005},
    {"manufacturer": "Jinko Solar",    "model": "Tiger Neo",
     "efficiency_percent": 22.53, "cost_per_wp_usd": 0.45, "temp_coeff_pct_per_C": -0.29,
     "voc_v": 39.38, "isc_a": 13.86, "vmp_v": 32.81, "imp_a": 13.41,
     "panel_power_w": 440, "panel_area_m2": 1.95,
     "cells_in_series": 54, "cells_in_parallel": 2, "degradation_rate": 0.004},
    {"manufacturer": "LONGi Solar",    "model": "Hi-MO 6",
     "efficiency_percent": 22.3,  "cost_per_wp_usd": 0.3,  "temp_coeff_pct_per_C": -0.29,
     "voc_v": 39.33, "isc_a": 14.22, "vmp_v": 33.04, "imp_a": 13.17,
     "panel_power_w": 435, "panel_area_m2": 1.95,
     "cells_in_series": 54, "cells_in_parallel": 2, "degradation_rate": 0.004},
    {"manufacturer": "Maxeon Solar",   "model": "Maxeon 7",
     "efficiency_percent": 24.1,  "cost_per_wp_usd": 3.5,  "temp_coeff_pct_per_C": -0.27,
     "voc_v": 83.0,  "isc_a": 6.6,  "vmp_v": 71.4,  "imp_a": 6.23,
     "panel_power_w": 445, "panel_area_m2": 1.85,
     "cells_in_series": 112, "cells_in_parallel": 1, "degradation_rate": 0.0025},
    {"manufacturer": "Aiko Solar",     "model": "Neostar 2P",
     "efficiency_percent": 24.3,  "cost_per_wp_usd": 0.55, "temp_coeff_pct_per_C": -0.26,
     "voc_v": 41.36, "isc_a": 14.41, "vmp_v": 34.92, "imp_a": 13.9,
     "panel_power_w": 485, "panel_area_m2": 1.99,
     "cells_in_series": 54, "cells_in_parallel": 2, "degradation_rate": 0.0035},
]

BATTERY_CATALOG = [
    {"manufacturer": "Tesla",     "model": "Powerwall 3",
     "usable_capacity_kwh": 13.5, "nominal_voltage_v": "52 - 92 (DC)",
     "max_charge_power_kw": 5.0,  "max_discharge_power_kw": 11.5,
     "round_trip_efficiency_pct": 97.5, "cycle_life": None,
     "cost_usd": 15400, "annual_degradation_rate": 0.02},
    {"manufacturer": "Enphase",   "model": "IQ Battery 5P",
     "usable_capacity_kwh": 5.0,  "nominal_voltage_v": "48.0",
     "max_charge_power_kw": 3.84, "max_discharge_power_kw": 3.84,
     "round_trip_efficiency_pct": 96.0, "cycle_life": 6000,
     "cost_usd": 8500, "annual_degradation_rate": 0.015},
    {"manufacturer": "Generac",   "model": "PWRcell M6",
     "usable_capacity_kwh": 18.0, "nominal_voltage_v": "360 - 420",
     "max_charge_power_kw": 10.5, "max_discharge_power_kw": 10.5,
     "round_trip_efficiency_pct": 96.5, "cycle_life": None,
     "cost_usd": 11500, "annual_degradation_rate": 0.02}
]

# BATTERY_CATALOG = [
#     {"manufacturer": "Tesla",     "model": "Powerwall 3",
#      "usable_capacity_kwh": 13.5, "nominal_voltage_v": "52 - 92 (DC)",
#      "max_charge_power_kw": 5.0,  "max_discharge_power_kw": 11.5,
#      "round_trip_efficiency_pct": 97.5, "cycle_life": None,
#      "cost_usd": 15400, "annual_degradation_rate": 0.02},
#     {"manufacturer": "Enphase",   "model": "IQ Battery 5P",
#      "usable_capacity_kwh": 5.0,  "nominal_voltage_v": "48.0",
#      "max_charge_power_kw": 3.84, "max_discharge_power_kw": 3.84,
#      "round_trip_efficiency_pct": 96.0, "cycle_life": 6000,
#      "cost_usd": 8500, "annual_degradation_rate": 0.015},
#     {"manufacturer": "Generac",   "model": "PWRcell M6",
#      "usable_capacity_kwh": 18.0, "nominal_voltage_v": "360 - 420",
#      "max_charge_power_kw": 10.5, "max_discharge_power_kw": 10.5,
#      "round_trip_efficiency_pct": 96.5, "cycle_life": None,
#      "cost_usd": 11500, "annual_degradation_rate": 0.02},
#     {"manufacturer": "SolarEdge", "model": "Home Battery",
#      "usable_capacity_kwh": 4.6,  "nominal_voltage_v": "44.8 - 56.5",
#      "max_charge_power_kw": 2.825,"max_discharge_power_kw": 4.096,
#      "round_trip_efficiency_pct": 94.5, "cycle_life": None,
#      "cost_usd": 12500, "annual_degradation_rate": 0.03},
#     {"manufacturer": "Panasonic", "model": "EverVolt H",
#      "usable_capacity_kwh": 13.5, "nominal_voltage_v": "153.6",
#      "max_charge_power_kw": 8.3,  "max_discharge_power_kw": 8.3,
#      "round_trip_efficiency_pct": 94.0, "cycle_life": 6000,
#      "cost_usd": 17500, "annual_degradation_rate": 0.025},
# ]

PR_PERFORMANCE_RATIO    = 0.80
INSTALLATION_COST_RATE  = 0.10
FEDERAL_ITC_RATE        = 0.30
UTILITY_INFLATION_RATE  = 0.06
DISCOUNT_RATE           = 0.07
O_AND_M_COST_PER_W_YR   = 0.005
NEM_EXPORT_CREDIT       = 0.10
INVERTER_REPLACEMENT_YR = 10
INVERTER_REPLACEMENT_USD= 2000
ANALYSIS_YEARS          = 10
SDGE_DAILY_FIXED_FEE    = 0.345
SDGE_TOTAL_CUSTOMERS    = 1_040_149
COASTAL_LON_REF         = -117.25
CITY_CENTER_LAT         = 32.7157
CITY_CENTER_LON         = -117.1611
MILP_SOC_MIN_FRAC = 0.10
MILP_SOC_INIT_FRAC= 0.50
EV_CHARGE_START_HOUR = 22
EV_CHARGE_END_HOUR   = 6
EV_CHARGER_POWER_KW  = 7.2
EV_DAILY_ENERGY_KWH  = 14.0
OPEN_METEO_URL = "https://archive-api.open-meteo.com/v1/archive"
_r = DISCOUNT_RATE; _n = ANALYSIS_YEARS
ANNUITY_FACTOR = _r*(1+_r)**_n / ((1+_r)**_n - 1)
PV_MULTIPLIER = sum(
    (1 + UTILITY_INFLATION_RATE)**(y - 1) / (1 + DISCOUNT_RATE)**y
    for y in range(1, ANALYSIS_YEARS + 1)
)

# ── Battery option enumeration ─────────────────────────────────────────────
# Pre-enumerate all (model × units) combinations so the MILP selects whole
# catalog units directly via binary variables — no floor/ceil post-processing.
# MAX_BATTERY_UNITS: maximum stacks of any single model allowed.
MAX_BATTERY_UNITS = 4
# Option 0 = no battery; options 1..K-1 = each (model, units) pair
BATTERY_OPTIONS = [
    {"label": "none", "cap_kwh": 0.0, "max_charge_kw": 0.0,
     "max_discharge_kw": 0.0, "cost_usd": 0, "bat_obj": None, "units": 0}
]
for _bat in BATTERY_CATALOG:
    for _u in range(1, MAX_BATTERY_UNITS + 1):
        BATTERY_OPTIONS.append({
            "label":            f"{_u}x {_bat['manufacturer']} {_bat['model']}",
            "cap_kwh":          _bat["usable_capacity_kwh"]     * _u,
            "max_charge_kw":    _bat["max_charge_power_kw"]     * _u,
            "max_discharge_kw": _bat["max_discharge_power_kw"]  * _u,
            "cost_usd":         _bat["cost_usd"]                * _u,
            "bat_obj":          _bat,
            "units":            _u,
        })
# K = 21: option[0]=none, options[1..20] = 5 models × 4 unit counts


# ══════════════════════════════════════════════════════════════════════════════
# CACHES
# ══════════════════════════════════════════════════════════════════════════════

_irr_cache   = {}
_tariff_cache= {}
_eia_cache   = {}


# ══════════════════════════════════════════════════════════════════════════════
# HELPER FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════

def _irradiance_shape_factor(hour_of_day, day_of_year):
    sunrise = 6.0
    sunset  = 18.0 + 2.0 * np.sin(2.0 * np.pi * (day_of_year - 80) / 365.0)
    if hour_of_day < sunrise or hour_of_day >= sunset:
        return 0.0
    return max(0.0, np.sin(np.pi * (hour_of_day - sunrise) / (sunset - sunrise)))


def _fetch_irradiance(latitude, longitude):
    key = (round(latitude, 4), round(longitude, 4))
    if key in _irr_cache:
        print(f"    [IRR] Cache hit for {key}")
        return _irr_cache[key]
    try:
        end    = datetime.now() - timedelta(days=7)
        start  = end.replace(year=end.year - 5)
        params = {"latitude": latitude, "longitude": longitude,
                  "start_date": start.strftime("%Y-%m-%d"),
                  "end_date":   end.strftime("%Y-%m-%d"),
                  "daily": "shortwave_radiation_sum", "timezone": "auto"}
        r = requests.get(OPEN_METEO_URL, params=params, timeout=60)
        r.raise_for_status()
        vals = [v for v in r.json()["daily"]["shortwave_radiation_sum"] if v is not None]
        if not vals:
            raise ValueError("empty API response")
        result = round(np.mean(vals) / 3.6 * 365, 1)
        print(f"    [IRR] Fetched {result} kWh/m²/yr for {key}")
    except Exception as e:
        print(f"    [IRR WARNING] API failed ({e}). Using SD default 2080.")
        result = 2080.0
    _irr_cache[key] = result
    return result


def _build_hourly_tariffs(rate_plan, year=2024):
    if rate_plan in _tariff_cache:
        return _tariff_cache[rate_plan]
    if rate_plan in ("TOU_DR", "TOU_DR1"):
        path = TOU_DR_PATH if rate_plan == "TOU_DR" else TOU_DR1_PATH
        df   = pd.read_csv(path, parse_dates=["date"])
        lk   = df[df["date"].dt.year == year].set_index("date")[
               ["on_peak_$/kwh", "off_peak_$/kwh", "super_off_peak_$/kwh"]]
        use_daily = True
    elif rate_plan == "TOU_DR2":
        df = pd.read_csv(TOU_DR2_PATH, parse_dates=["start_date", "end_date"])
        lk = df[df["year"] == year].copy()
        use_daily = False
    else:
        raise ValueError(f"Unknown rate_plan '{rate_plan}'")
    out = []
    y0  = datetime(year, 1, 1)
    for h in range(8760):
        dt     = y0 + timedelta(hours=h)
        hr, mo = dt.hour, dt.month
        on_p   = (16 <= hr < 21)
        sup_p  = (hr < 6) or (mo in (3, 4) and 10 <= hr < 14)
        if use_daily:
            dk = pd.Timestamp(dt.date())
            if dk not in lk.index:
                dk = min(lk.index, key=lambda d: abs((d - dk).days))
            row = lk.loc[dk]
            out.append(float(row["on_peak_$/kwh"] if on_p else
                             (row["super_off_peak_$/kwh"] if sup_p else row["off_peak_$/kwh"])))
        else:
            is_sum = (6 <= mo <= 10)
            ts = pd.Timestamp(dt.date()); seg = None
            for _, s in lk.iterrows():
                if s["start_date"] <= ts <= s["end_date"]:
                    seg = s; break
            if seg is None:
                out.append(0.38); continue
            col = (("summer_on_peak_$/kwh" if is_sum else "winter_on_peak_$/kwh") if on_p
                   else ("summer_off_peak_$/kwh" if is_sum else "winter_off_peak_$/kwh"))
            out.append(float(seg[col]))
    _tariff_cache[rate_plan] = out
    return out


def _load_household_profile_from_eia(latitude, longitude,
                                      annual_kwh_override=None,
                                      num_evs=0, num_people=3,
                                      num_daytime_occupants=1):
    cache_key = (round(latitude,4), round(longitude,4),
                 annual_kwh_override, num_evs, num_people, num_daytime_occupants)
    if cache_key in _eia_cache:
        return _eia_cache[cache_key]

    if not os.path.exists(EIA_LOAD_PATH):
        raise FileNotFoundError(f"EIA file not found: {EIA_LOAD_PATH}")

    df = pd.read_csv(EIA_LOAD_PATH)
    df["dt_utc"]   = pd.to_datetime(df["Timestamp_UTC"])
    df["dt_local"] = df["dt_utc"].dt.tz_localize("UTC").dt.tz_convert("America/Los_Angeles")
    df["kw"]       = df["MW_Load"] * 1000.0 / SDGE_TOTAL_CUSTOMERS
    df["year"]     = df["dt_local"].dt.year
    df["doy"]      = df["dt_local"].dt.dayofyear
    df["hour"]     = df["dt_local"].dt.hour
    df["hoy"]      = ((df["doy"] - 1) * 24 + df["hour"]).clip(0, 8759)

    full_years = df["year"].value_counts()
    full_years = full_years[full_years >= 8000].index.tolist()
    df_full    = df[df["year"].isin(full_years)]
    avg_hoy    = (df_full.groupby("hoy")["kw"].mean()
                         .reindex(range(8760)).interpolate(method="linear")
                         .ffill().bfill())
    profile = avg_hoy.values.copy()

    loc_seed = int(hashlib.sha256(f"{latitude}_{longitude}".encode()).hexdigest(), 16) % (2**32)
    rng_sol  = np.random.RandomState(loc_seed + 1000)
    rng_mg   = np.random.RandomState(loc_seed + 3000)

    dc = longitude - COASTAL_LON_REF
    if   dc >= 0.15:  f1 = 1.25
    elif dc >= 0.10:  f1 = 1.05 + (dc - 0.10)*4.0
    elif dc >= 0:     f1 = 0.95 + dc*1.0
    elif dc >= -0.05: f1 = 0.90 + (dc + 0.05)*1.0
    else:             f1 = 0.85

    if   latitude < 32.60: f2 = 1.10
    elif latitude < 32.70: f2 = 1.05
    elif latitude < 32.85: f2 = 1.00
    elif latitude < 32.95: f2 = 0.95
    else:                  f2 = 0.90

    f3 = 1.0 + (max(0, longitude - COASTAL_LON_REF) + max(0, latitude - 32.70)*2.0)*0.15
    f4 = 0.70 + 0.10 * min(num_people, 6)

    dist_c = math.sqrt((latitude - CITY_CENTER_LAT)**2 + (longitude - CITY_CENTER_LON)**2)
    f5 = (0.7 if dist_c < 0.03 else 0.9 if dist_c < 0.08 else 1.1 if dist_c < 0.15 else 1.3)

    is_coastal    = longitude < -117.20
    is_north      = latitude  >  32.80
    is_urban_core = dist_c    <   0.05
    if   is_coastal and is_north: f6 = 1.15
    elif is_coastal:              f6 = 1.05
    elif is_urban_core:           f6 = 1.25
    elif longitude > -117.00:     f6 = 0.95
    else:                         f6 = 1.10

    is_south = latitude < 32.75; is_urban = dist_c < 0.10
    mg_prob  = 0.25 if (is_south and is_urban) else (0.15 if is_urban else 0.10)
    f9 = rng_mg.uniform(1.20, 1.50) if rng_mg.random() < mg_prob else 1.0

    profile *= f1 * f2 * f3 * f4 * f5 * f6 * f9

    is_affluent = is_coastal and is_north
    sol_prob = (0.35 if is_affluent else 0.20 if (is_coastal or longitude > -117.00)
                else 0.05 if is_urban_core else 0.15)
    if rng_sol.random() < sol_prob:
        hrs       = np.arange(8760) % 24
        intensity = np.clip(np.sin((hrs - 6)*np.pi/12), 0, 1)
        intensity[(hrs < 6)|(hrs > 18)] = 0
        max_red   = rng_sol.uniform(0.4, 0.7)
        profile  *= 1.0 - intensity*(1.0 - max_red)

    if num_evs > 0:
        hrs = np.arange(8760) % 24
        hours_needed = int(np.ceil((EV_DAILY_ENERGY_KWH * num_evs) / EV_CHARGER_POWER_KW))
        end_h = (EV_CHARGE_START_HOUR + hours_needed) % 24
        is_chg = ((hrs >= EV_CHARGE_START_HOUR)|(hrs < end_h)) \
                 if EV_CHARGE_START_HOUR >= end_h \
                 else ((hrs >= EV_CHARGE_START_HOUR)&(hrs < end_h))
        profile += np.where(is_chg, EV_CHARGER_POWER_KW * num_evs, 0.0)

    if num_daytime_occupants > 0:
        hrs = np.arange(8760) % 24
        profile = np.where((hrs >= 9)&(hrs < 17),
                           profile * (1.0 + 0.05 * num_daytime_occupants), profile)

    profile *= np.random.RandomState(loc_seed).normal(1.0, 0.03, size=8760)
    profile  = np.clip(profile, 0, None)

    prof_sum = float(profile.sum())
    if annual_kwh_override is not None and annual_kwh_override > 0 and prof_sum > 0:
        profile *= annual_kwh_override / prof_sum
        ann_kwh  = annual_kwh_override
    else:
        ann_kwh = round(prof_sum, 1)

    result = (profile.tolist(), ann_kwh)
    _eia_cache[cache_key] = result
    return result


def _build_annual_load_profile(annual_kwh, num_evs, num_people, num_daytime_occupants):
    shape = np.array([0.40,0.35,0.32,0.30,0.30,0.35,
                      0.50,0.70,0.85,0.80,0.75,0.72,
                      0.70,0.68,0.65,0.65,0.70,0.90,
                      1.00,0.95,0.85,0.75,0.65,0.50])
    ev24 = np.zeros(24)
    if num_evs > 0:
        ch = (list(range(EV_CHARGE_START_HOUR, 24)) + list(range(0, EV_CHARGE_END_HOUR)))
        ch = ch[:int(np.ceil(EV_DAILY_ENERGY_KWH / EV_CHARGER_POWER_KW))]
        for h in ch: ev24[h] = EV_CHARGER_POWER_KW * num_evs
    ev_ann = ev24.sum() * 365
    y0 = datetime(2024, 1, 1)
    raw = []
    for h in range(8760):
        dt = y0 + timedelta(hours=h)
        sm = 1.20 if 6<=dt.month<=10 else (0.95 if dt.month in (3,4,5,11) else 1.00)
        om = 1.0 + 0.05*num_daytime_occupants if 9<=dt.hour<17 else 1.0
        ps = 0.70 + 0.10*min(num_people, 6)
        raw.append(shape[dt.hour]*sm*om*ps)
    arr   = np.array(raw)
    scale = max(annual_kwh - ev_ann, 0.0)/arr.sum() if arr.sum() > 0 else 1.0
    return [float(raw[h]*scale + ev24[h%24]) for h in range(8760)]


def _select_panel(panel_brand):
    if panel_brand is None or str(panel_brand).strip().lower() in ("none",""):
        return max(SOLAR_PANEL_CATALOG, key=lambda p: p["efficiency_percent"]/p["cost_per_wp_usd"])
    brand = str(panel_brand).strip().lower()
    m = [p for p in SOLAR_PANEL_CATALOG if brand in p["manufacturer"].lower()]
    return m[0] if m else max(SOLAR_PANEL_CATALOG,
                              key=lambda p: p["efficiency_percent"]/p["cost_per_wp_usd"])


def _build_hourly_pv_output(panel, n_panels, irr_kwh_m2_yr):
    pkw = panel["panel_power_w"]/1000.0
    raw = [n_panels*pkw*_irradiance_shape_factor(h%24,(h//24)%365+1)*PR_PERFORMANCE_RATIO
           for h in range(8760)]
    tgt = n_panels*pkw*irr_kwh_m2_yr*PR_PERFORMANCE_RATIO
    tot = sum(raw)
    if tot > 0:
        raw = [v*tgt/tot for v in raw]
    return raw


def _run_dispatch_simulation(hourly_load_kw, hourly_pv_kw, hourly_tariffs, battery):
    has_b = battery is not None
    cap   = battery["usable_capacity_kwh"]          if has_b else 0.0
    maxc  = battery["max_charge_power_kw"]           if has_b else 0.0
    maxd  = battery["max_discharge_power_kw"]        if has_b else 0.0
    eta   = battery["round_trip_efficiency_pct"]/100.0 if has_b else 1.0
    soc   = cap * 0.5
    tot_imp = tot_exp = imp_cost = exp_cred = 0.0
    for h in range(8760):
        net = hourly_load_kw[h] - hourly_pv_kw[h]
        if net < 0:
            sur  = -net
            ckw  = min(min(maxc, (cap - soc) / eta) if has_b else 0.0, sur)
            soc += ckw * eta; export = sur - ckw; gbuy = 0.0
        else:
            dkw  = min(maxd, soc) if has_b else 0.0
            dkw  = min(net, dkw)
            soc -= dkw; gbuy = net - dkw; export = 0.0
        soc = max(0.0, min(cap, soc))
        tot_imp += gbuy; tot_exp += export
        imp_cost += gbuy   * hourly_tariffs[h]
        exp_cred += export * NEM_EXPORT_CREDIT
    return {"annual_import_kwh": round(tot_imp, 1),
            "annual_export_kwh": round(tot_exp, 1),
            "import_cost_usd":   round(imp_cost, 2),
            "export_credit_usd": round(exp_cred, 2)}


def _compute_economics(dispatch, panel, n_panels, battery, battery_units,
                        annual_load_kwh, avg_tariff, with_battery):
    array_w  = n_panels * panel["panel_power_w"]
    pv_cost  = array_w * panel["cost_per_wp_usd"]
    bat_cost = (battery["cost_usd"] * battery_units) if battery else 0.0
    gross    = (pv_cost + bat_cost) * (1 + INSTALLATION_COST_RATE)
    net_cap  = gross * (1 - FEDERAL_ITC_RATE)
    fixed_ann= SDGE_DAILY_FIXED_FEE * 365
    trad_y1  = annual_load_kwh * avg_tariff + fixed_ann
    solar_y1 = dispatch["import_cost_usd"] - dispatch["export_credit_usd"] + fixed_ann
    eff_import_rate = (dispatch["import_cost_usd"] / dispatch["annual_import_kwh"]
                       if dispatch["annual_import_kwh"] > 0 else avg_tariff)

    cumul = 0.0; payback = None; npv = -net_cap
    for y in range(1, ANALYSIS_YEARS + 1):
        inf  = (1 + UTILITY_INFLATION_RATE)**(y - 1)
        disc = (1 + DISCOUNT_RATE)**y
        trad = trad_y1 * inf
        scl  = (1 - panel["degradation_rate"])**(y - 1)
        imp  = annual_load_kwh - (annual_load_kwh - dispatch["annual_import_kwh"]) * scl
        exp  = dispatch["annual_export_kwh"] * scl
        sol  = imp * eff_import_rate * inf - exp * NEM_EXPORT_CREDIT + fixed_ann
        om   = O_AND_M_COST_PER_W_YR * array_w * inf
        inv  = INVERTER_REPLACEMENT_USD if y == INVERTER_REPLACEMENT_YR else 0.0
        sav  = trad - (sol + om + inv)
        cumul += sav
        if payback is None and cumul >= net_cap: payback = y
        npv += sav / disc
    sav_y1 = trad_y1 - solar_y1 - O_AND_M_COST_PER_W_YR * array_w
    return {"gross_capex_usd":         round(gross,   2),
            "net_capex_after_itc_usd": round(net_cap, 2),
            "annual_savings_usd":      round(sav_y1,  2),
            "simple_payback_years":    payback if payback else float("inf"),
            "npv_usd":                 round(npv, 2),
            "annual_import_kwh":       dispatch["annual_import_kwh"],
            "annual_export_kwh":       dispatch["annual_export_kwh"]}


# ══════════════════════════════════════════════════════════════════════════════
# MILP OPTIMIZER — binary battery option selection
# ══════════════════════════════════════════════════════════════════════════════

def optimize_energy_system_milp(panel, irr_kwh_m2_yr, hourly_load_kw,
                                 hourly_tariffs, roof_area_m2, budget_usd,
                                 force_battery, annual_kwh):
    """
    MILP that jointly optimises panel count (integer) and battery choice
    (binary one-hot over BATTERY_OPTIONS catalog entries).

    Variable layout:
      I_P              = 0              integer: panel count
      I_Y[0..K-1]      = 1..K           binary: one-hot battery option selector
      I_CAP            = K+1            continuous: effective capacity (kWh)
      I_MAXCHG         = K+2            continuous: max charge power (kW)
      I_MAXDIS         = K+3            continuous: max discharge power (kW)
      I_IMP[0..T-1]    = K+4..K+T+3    continuous: hourly grid import (kW)
      I_EXP[0..T-1]    = K+T+4..K+2T+3
      I_CHG[0..T-1]    = K+2T+4..K+3T+3
      I_DIS[0..T-1]    = K+3T+4..K+4T+3
      I_SOC[0..T]      = K+4T+4..K+5T+4  (T+1 values)
    """
    T          = len(hourly_load_kw)
    panel_kw   = panel["panel_power_w"] / 1000.0
    panel_area = panel["panel_area_m2"]
    panel_cost = panel["panel_power_w"] * panel["cost_per_wp_usd"]
    max_panels = int(roof_area_m2 / panel_area)

    raw_shapes = np.array([_irradiance_shape_factor(h % 24, (h // 24) % 365 + 1)
                           for h in range(T)])
    raw_sum    = raw_shapes.sum()
    target_sum = irr_kwh_m2_yr * PR_PERFORMANCE_RATIO
    irr_h      = raw_shapes * (target_sum / raw_sum) if raw_sum > 0 else raw_shapes

    Load = np.array(hourly_load_kw, dtype=float)
    Pbuy = np.array(hourly_tariffs,  dtype=float)

    K = len(BATTERY_OPTIONS)

    # Per-option eta: "none" option gets eta=1.0 (irrelevant, capacity=0)
    ETA = np.array([
        (opt["bat_obj"]["round_trip_efficiency_pct"] / 100.0
         if opt["bat_obj"] is not None else 1.0)
        for opt in BATTERY_OPTIONS
    ])

    # ── Variable indices ─────────────────────────────────────────────────
    # Disaggregated layout: separate CHG[k,t] and DIS[k,t] per battery option.
    # Big-M constraints enforce CHG[k,t]=DIS[k,t]=0 when y[k]=0, so SOC
    # dynamics can use each option's exact eta without bilinear terms.
    #
    #   I_P               = 0                integer: panel count
    #   I_Y[0..K-1]       = 1..K             binary: one-hot battery selector
    #   I_CAP             = K+1              continuous: effective capacity (kWh)
    #   I_MAXCHG          = K+2              continuous: max charge power (kW)
    #   I_MAXDIS          = K+3              continuous: max discharge power (kW)
    #   I_IMP[0..T-1]     = K+4..K+T+3      continuous: hourly grid import (kW)
    #   I_EXP[0..T-1]     = K+T+4..K+2T+3
    #   I_CHG_k[k,0..T-1] = K+2T+4 + k*T   continuous: charge power for option k
    #   I_DIS_k[k,0..T-1] = K+(K+2)*T+4+k*T
    #   I_SOC[0..T]       = K+2*(K+1)*T+4   (T+1 values)
    I_P      = 0
    I_Y      = np.arange(1, K + 1)
    I_CAP    = K + 1
    I_MAXCHG = K + 2
    I_MAXDIS = K + 3
    off      = K + 4
    I_IMP    = np.arange(off,         off + T)
    I_EXP    = np.arange(off + T,     off + 2*T)
    off2     = off + 2*T
    # I_CHG_k[k] is a length-T array of variable indices for option k's charge
    I_CHG_k  = np.array([np.arange(off2 + k*T, off2 + (k+1)*T) for k in range(K)])
    off3     = off2 + K*T
    # I_DIS_k[k] is a length-T array of variable indices for option k's discharge
    I_DIS_k  = np.array([np.arange(off3 + k*T, off3 + (k+1)*T) for k in range(K)])
    off4     = off3 + K*T
    I_SOC    = np.arange(off4, off4 + T + 1)   # T+1 elements
    n_vars   = off4 + T + 1

    # Big-M for charge/discharge: upper bound is the max power any option can deliver
    M_pow = max(max(opt["max_charge_kw"], opt["max_discharge_kw"])
                for opt in BATTERY_OPTIONS)

    # ── Objective ────────────────────────────────────────────────────────
    c_obj = np.zeros(n_vars)
    # Panel annualised capex
    c_obj[I_P] = panel_cost * (1 + INSTALLATION_COST_RATE) * (1 - FEDERAL_ITC_RATE) * ANNUITY_FACTOR
    # Battery annualised capex — each option priced at its true catalog cost
    for k, opt in enumerate(BATTERY_OPTIONS):
        c_obj[I_Y[k]] = (opt["cost_usd"] * (1 + INSTALLATION_COST_RATE)
                         * (1 - FEDERAL_ITC_RATE) * ANNUITY_FACTOR)
    # Operating costs (inflation-adjusted NPV over horizon)
    c_obj[I_IMP] =  Pbuy * PV_MULTIPLIER
    c_obj[I_EXP] = -NEM_EXPORT_CREDIT * PV_MULTIPLIER

    # ── Bounds ───────────────────────────────────────────────────────────
    lb = np.zeros(n_vars)
    ub = np.full(n_vars, np.inf)
    ub[I_P]      = max_panels
    ub[I_Y]      = 1.0          # binary: 0 ≤ y[k] ≤ 1
    ub[I_CAP]    = sum(opt["cap_kwh"]          for opt in BATTERY_OPTIONS)
    ub[I_MAXCHG] = sum(opt["max_charge_kw"]    for opt in BATTERY_OPTIONS)
    ub[I_MAXDIS] = sum(opt["max_discharge_kw"] for opt in BATTERY_OPTIONS)
    # Per-option CHG/DIS bounded by their catalog max (tightens big-M)
    for k, opt in enumerate(BATTERY_OPTIONS):
        ub[I_CHG_k[k]] = opt["max_charge_kw"]
        ub[I_DIS_k[k]] = opt["max_discharge_kw"]
    ub[I_SOC]    = ub[I_CAP]
    if force_battery:
        ub[I_Y[0]] = 0.0        # forbid "none" option → forces a real battery

    # ── Integrality ──────────────────────────────────────────────────────
    integrality = np.zeros(n_vars)   # 0 = continuous
    integrality[I_P]   = 1           # integer
    integrality[I_Y]   = 1           # binary (enforced by 0/1 bounds)

    # ── Constraint builder ───────────────────────────────────────────────
    rs = []; cs = []; ds = []; lo_c = []; hi_c = []; ridx = [0]

    def _add(entries, lo, hi):
        for col, val in entries:
            rs.append(ridx[0]); cs.append(col); ds.append(val)
        lo_c.append(lo); hi_c.append(hi); ridx[0] += 1

    def eq(e, r): _add(e, r, r)
    def le(e, r): _add(e, -np.inf, r)

    # (1) Exactly one battery option selected
    eq([(I_Y[k], 1.0) for k in range(K)], 1.0)

    # (2) Link scalar capacity / power variables to chosen option
    eq([(I_CAP, 1.0)]    + [(I_Y[k], -BATTERY_OPTIONS[k]["cap_kwh"])         for k in range(K)], 0.0)
    eq([(I_MAXCHG, 1.0)] + [(I_Y[k], -BATTERY_OPTIONS[k]["max_charge_kw"])   for k in range(K)], 0.0)
    eq([(I_MAXDIS, 1.0)] + [(I_Y[k], -BATTERY_OPTIONS[k]["max_discharge_kw"]) for k in range(K)], 0.0)

    # (3) Hourly energy balance — charge/discharge summed across all options
    #     (only the chosen option's variables will be non-zero via big-M below)
    for t in range(T):
        pvc = panel_kw * irr_h[t]
        entries = [(I_P, pvc), (I_IMP[t], 1.0), (I_EXP[t], -1.0)]
        for k in range(K):
            entries += [(I_CHG_k[k][t], -1.0), (I_DIS_k[k][t], 1.0)]
        eq(entries, Load[t])

    # (4) SOC dynamics using each option's exact eta
    #     SOC[t+1] = SOC[t] + eta[k]*CHG[k,t] - DIS[k,t]/eta[k]
    #     Rearranged: -SOC[t] + SOC[t+1] - sum_k(eta[k]*CHG[k,t]) + sum_k(DIS[k,t]/eta[k]) = 0
    for t in range(T):
        entries = [(I_SOC[t], -1.0), (I_SOC[t + 1], 1.0)]
        for k in range(K):
            entries += [(I_CHG_k[k][t], -ETA[k]), (I_DIS_k[k][t], 1.0 / ETA[k])]
        eq(entries, 0.0)

    # (5) Initial SOC
    eq([(I_SOC[0], 1.0), (I_CAP, -MILP_SOC_INIT_FRAC)], 0.0)

    # (6) SOC bounds
    for t in range(T + 1):
        le([(I_SOC[t],  1.0), (I_CAP, -1.0)],              0.0)  # SOC ≤ CAP
        le([(I_SOC[t], -1.0), (I_CAP,  MILP_SOC_MIN_FRAC)], 0.0)  # SOC ≥ CAP*min_frac

    # (7) Charge / discharge power limits + big-M: force CHG[k,t]=DIS[k,t]=0
    #     when y[k]=0. Upper bounds already cap each at catalog max; big-M here
    #     drives them to zero for unchosen options.
    for k in range(K):
        for t in range(T):
            le([(I_CHG_k[k][t], 1.0), (I_Y[k], -M_pow)], 0.0)
            le([(I_DIS_k[k][t], 1.0), (I_Y[k], -M_pow)], 0.0)

    # (8) Export ≤ PV generation
    for t in range(T):
        le([(I_EXP[t], 1.0), (I_P, -panel_kw * irr_h[t])], 0.0)

    # (9) Roof area
    le([(I_P, panel_area)], roof_area_m2)

    # (10) Budget: panels + chosen battery option (exact catalog cost)
    budget_entries = [(I_P, panel_cost * (1 + INSTALLATION_COST_RATE))]
    for k, opt in enumerate(BATTERY_OPTIONS):
        budget_entries.append((I_Y[k], opt["cost_usd"] * (1 + INSTALLATION_COST_RATE)))
    le(budget_entries, budget_usd)

    # ── Solve ────────────────────────────────────────────────────────────
    n_con = ridx[0]
    A_sp  = csc_matrix((np.array(ds), (np.array(rs), np.array(cs))),
                       shape=(n_con, n_vars))
    con   = LinearConstraint(A_sp, lb=np.array(lo_c), ub=np.array(hi_c))

    result = milp(c=c_obj, constraints=con, integrality=integrality,
                  bounds=Bounds(lb=lb, ub=ub),
                  options={"disp": False, "time_limit": 300, "mip_rel_gap": 0.005})

    STATUS = {0: "Optimal", 1: "Iteration limit", 2: "Infeasible", 3: "Unbounded",
              4: "Infeasible or unbounded", 5: "Numerical error",
              6: "Time limit — best feasible"}

    if result.status in (0, 6) and result.x is not None:
        x = result.x
        n_panels_milp = max(1, int(round(x[I_P])))
        # Identify which battery option was chosen (highest y value ≈ 1.0)
        chosen_k   = int(np.argmax(x[I_Y]))
        chosen_opt = BATTERY_OPTIONS[chosen_k]
        return {
            "n_panels":            n_panels_milp,
            "bat_option":          chosen_opt,      # full option dict
            "status":              STATUS.get(result.status, f"Code {result.status}"),
            "annual_import_kwh":   round(float(x[I_IMP].sum()), 1),
            "annual_export_kwh":   round(float(x[I_EXP].sum()), 1),
        }
    else:
        return {
            "n_panels":            max(1, int(roof_area_m2 // panel_area)),
            "bat_option":          BATTERY_OPTIONS[0],   # fallback: no battery
            "status":              f"Fallback ({STATUS.get(result.status, result.status)})",
            "annual_import_kwh":   round(float(Load.sum()), 1),
            "annual_export_kwh":   0.0,
        }


# ══════════════════════════════════════════════════════════════════════════════
# FULL PIPELINE PER ROW
# ══════════════════════════════════════════════════════════════════════════════

def _make_stacked_battery(bat_option):
    """
    Given a BATTERY_OPTIONS entry, return (bs, bunt, stk) ready for dispatch.
    bs   = raw catalog entry (or None)
    bunt = integer unit count
    stk  = scaled dict with aggregate capacity/power (or None)
    """
    bs   = bat_option["bat_obj"]
    bunt = bat_option["units"]
    if bs is None or bunt == 0:
        return None, 0, None
    stk = dict(bs)
    stk["usable_capacity_kwh"]    = bs["usable_capacity_kwh"]    * bunt
    stk["max_charge_power_kw"]    = bs["max_charge_power_kw"]    * bunt
    stk["max_discharge_power_kw"] = bs["max_discharge_power_kw"] * bunt
    return bs, bunt, stk


def run_optimizer_for_row(row_dict):
    """Runs the full MILP pipeline for one manifest row."""
    lat    = float(row_dict["latitude"])
    lon    = float(row_dict["longitude"])
    budget = float(row_dict["budget_usd"])
    roof_l = float(row_dict["roof_length_m"])
    roof_w = float(row_dict["roof_breadth_m"])
    rp     = str(row_dict["rate_plan"])
    br     = str(row_dict["panel_brand"]).strip()
    nevs   = int(row_dict["num_evs"])
    nppl   = int(row_dict["num_people"])
    nday   = int(row_dict["num_daytime_occupants"])

    # 1. Irradiance (cached by lat/lon)
    irr = _fetch_irradiance(lat, lon)

    # 2. Tariffs (cached by rate plan)
    tars = _build_hourly_tariffs(rp)
    avgt = float(np.mean(tars))

    # 3. Load profile — annual_consumption_kwh = None (auto from EIA)
    try:
        load, akwh = _load_household_profile_from_eia(lat, lon, None, nevs, nppl, nday)
    except FileNotFoundError:
        akwh = 7500.0
        load = _build_annual_load_profile(akwh, nevs, nppl, nday)

    # 4. Panel selection
    panel = _select_panel(None if br.lower() == "none" else br)

    # 5. True physical roof fit (Tetris — try both panel orientations)
    p_w = panel["panel_area_m2"]**0.5 * 0.75
    p_h = panel["panel_area_m2"] / p_w
    true_max = max(int(roof_w / p_w) * int(roof_l / p_h),
                   int(roof_w / p_h) * int(roof_l / p_w))
    effective_roof = true_max * panel["panel_area_m2"]

    # 6. MILP 1 — OPTIMAL (unlimited budget; solver picks panels + battery freely)
    o1 = optimize_energy_system_milp(panel, irr, load, tars,
                                     effective_roof, float("inf"), False, akwh)

    # 7. MILP 2 — RECOMMENDED (strict budget; solver picks panels + battery within budget)
    o2 = optimize_energy_system_milp(panel, irr, load, tars,
                                     effective_roof, budget, False, akwh)

    # 8. Dispatch simulation + 10-yr economics
    # MILP already chose whole catalog units — no floor/ceil needed.

    # OPT system
    bs1, bunt1, stk1 = _make_stacked_battery(o1["bat_option"])
    pv1 = _build_hourly_pv_output(panel, o1["n_panels"], irr)
    d1  = _run_dispatch_simulation(load, pv1, tars, stk1)
    ec1 = _compute_economics(d1, panel, o1["n_panels"], bs1, bunt1, akwh, avgt, bool(bs1))

    # REC system — MILP already respects budget for both panels and battery.
    # n_panels from MILP is already budget-feasible; no reinvestment step needed
    # because the MILP jointly optimised both dimensions simultaneously.
    bs2, bunt2, stk2 = _make_stacked_battery(o2["bat_option"])
    pv2 = _build_hourly_pv_output(panel, o2["n_panels"], irr)
    d2  = _run_dispatch_simulation(load, pv2, tars, stk2)
    ec2 = _compute_economics(d2, panel, o2["n_panels"], bs2, bunt2, akwh, avgt, bool(bs2))

    # 9. Cell spec
    ns  = panel["cells_in_series"]; np_ = panel["cells_in_parallel"]
    vc  = panel["vmp_v"] / ns
    ic  = (panel["panel_power_w"] / (ns * np_)) / vc if vc else 0

    # 10. String config helper
    def str_cfg(n):
        best_ns, best_np, best_diff = 1, n, 1e9
        for ns2 in range(1, n + 1):
            if n % ns2 != 0: continue
            diff = abs(ns2 * panel["vmp_v"] - 400)
            if diff < best_diff:
                best_diff = diff; best_ns = ns2; best_np = n // ns2
        return best_ns, best_np, round(best_ns * panel["vmp_v"], 2)

    opt_sns, opt_snp, opt_sv = str_cfg(max(1, o1["n_panels"]))
    rec_sns, rec_snp, rec_sv = str_cfg(max(1, o2["n_panels"]))

    max_bud_panels = min(true_max,
                         int(budget / (panel["panel_power_w"] * panel["cost_per_wp_usd"]
                                       * (1 + INSTALLATION_COST_RATE))))

    rec_bat_label = (f"{bs2['manufacturer']} {bs2['model']} ×{bunt2}"
                     if bs2 else "None")

    return {
        # Panel
        "panel_name":        f"{panel['manufacturer']} {panel['model']}",
        "panel_power_w":     panel["panel_power_w"],
        "panel_eff_pct":     panel["efficiency_percent"],
        "panel_area_m2":     panel["panel_area_m2"],
        "cell_vmp_v":        round(vc, 4),
        "cell_imp_a":        round(ic, 4),
        "max_panels_roof":   true_max,
        "max_panels_budget": max_bud_panels,
        "prod_per_panel":    round(panel["panel_power_w"] / 1000.0 * irr * PR_PERFORMANCE_RATIO, 1),
        "annual_kwh_auto":   None,
        "irradiance":        irr,
        "annual_load_kwh":   round(akwh, 1),
        # Optimal
        "opt_panels":      o1["n_panels"],
        "opt_kw_dc":       round(o1["n_panels"] * panel["panel_power_w"] / 1000, 3),
        "opt_prod_kwh":    round(o1["n_panels"] * panel["panel_power_w"] / 1000 * irr * PR_PERFORMANCE_RATIO, 1),
        "opt_gross_capex": ec1["gross_capex_usd"],
        "opt_net_capex":   ec1["net_capex_after_itc_usd"],
        "opt_savings":     ec1["annual_savings_usd"],
        "opt_payback":     ec1["simple_payback_years"],
        "opt_npv":         ec1["npv_usd"],
        "opt_sv":          opt_sv, "opt_snp": opt_snp, "opt_sns": opt_sns,
        "opt_array_area":  round(o1["n_panels"] * panel["panel_area_m2"], 2),
        "opt_budget_bind": "No",
        "opt_status":      o1["status"],
        # Recommended
        "rec_panels":      o2["n_panels"],
        "rec_kw_dc":       round(o2["n_panels"] * panel["panel_power_w"] / 1000, 3),
        "rec_prod_kwh":    round(o2["n_panels"] * panel["panel_power_w"] / 1000 * irr * PR_PERFORMANCE_RATIO, 1),
        "rec_gross_capex": ec2["gross_capex_usd"],
        "rec_net_capex":   ec2["net_capex_after_itc_usd"],
        "rec_savings":     ec2["annual_savings_usd"],
        "rec_payback":     ec2["simple_payback_years"],
        "rec_npv":         ec2["npv_usd"],
        "rec_sv":          rec_sv, "rec_snp": rec_snp, "rec_sns": rec_sns,
        "rec_array_area":  round(o2["n_panels"] * panel["panel_area_m2"], 2),
        "rec_budget_bind": "Yes" if ec2["gross_capex_usd"] >= budget * 0.97 else "No",
        "rec_status":      o2["status"],
        "rec_bat_kwh":     o2["bat_option"]["cap_kwh"],      # actual deployed kWh
        "rec_bat_model":   rec_bat_label,
    }


# ══════════════════════════════════════════════════════════════════════════════
# TXT PARSER
# ══════════════════════════════════════════════════════════════════════════════

def _grab(text, label):
    pat = re.escape(label) + r"[:\s]+([0-9,\.\-\$]+)"
    m = re.search(pat, text, re.IGNORECASE)
    if not m: return None
    raw = re.sub(r"[,$\s]", "", m.group(1))
    try:    return float(raw)
    except: return None


def parse_txt(filepath):
    if not os.path.exists(filepath):
        return None
    with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
        raw = f.read()

    blocks = [b.strip() for b in re.split(r"={5,}", raw) if b.strip()]
    opt_block = rec_block = ""
    for i, blk in enumerate(blocks):
        if re.search(r"1\.\s*OPTIMAL SYSTEM",    blk, re.IGNORECASE) and i + 1 < len(blocks):
            opt_block = blocks[i + 1]
        if re.search(r"2\.\s*RECOMMENDED SYSTEM", blk, re.IGNORECASE) and i + 1 < len(blocks):
            rec_block = blocks[i + 1]

    def parse_block(blk):
        d = {}
        d["panels"]            = _grab(blk, "Panels")
        m_kw = re.search(r"System size \(kW DC\)[:\s]+([0-9\.]+)", blk, re.IGNORECASE)
        d["kw_dc"]             = float(m_kw.group(1)) if m_kw else None
        d["offset_pct"]        = _grab(blk, "Target offset")
        d["annual_prod_kwh"]   = _grab(blk, "Annual production")
        d["annual_savings_usd"]= _grab(blk, "Annual savings")
        d["capex_usd"]         = _grab(blk, "CAPEX estimate")
        d["payback_yr"]        = _grab(blk, "Payback period")
        m_bb = re.search(r"Budget binding\?.*?(Yes|No)", blk, re.IGNORECASE)
        d["budget_binding"]    = m_bb.group(1) if m_bb else None
        return d

    return {"optimal": parse_block(opt_block), "recommended": parse_block(rec_block)}


# ══════════════════════════════════════════════════════════════════════════════
# COLUMN DEFINITIONS
# ══════════════════════════════════════════════════════════════════════════════

INPUT_HEADERS = [
    "File #", "Location", "Latitude", "Longitude",
    "Budget (USD)", "Roof Area (m²)", "Roof Length (m)", "Roof Width (m)",
    "Rate Plan", "Panel Brand", "Num EVs", "Num People", "Num Daytime Occupants",
]
MATH_HEADERS = [
    "Selected Panel", "Panel Power (Wp)", "Panel Efficiency (%)", "Panel Area (m²)",
    "Cell Vmp (V)", "Cell Imp (A)",
    "Max Panels by Roof", "Max Panels by Budget",
    "Prod/Panel/yr (kWh)", "Annual Consumption kWh (None=auto)",
    "Irradiance (kWh/m²/yr)", "Annual Load kWh (EIA)",
]
OPT_MATH_HEADERS = [
    "OPT_MATH: Panels", "OPT_MATH: System kW DC", "OPT_MATH: Annual Prod (kWh)",
    "OPT_MATH: Gross CAPEX ($)", "OPT_MATH: Net CAPEX ITC ($)",
    "OPT_MATH: Annual Savings ($)", "OPT_MATH: Payback (yr)", "OPT_MATH: NPV ($)",
    "OPT_MATH: String Voltage (V)", "OPT_MATH: Parallel Strings",
    "OPT_MATH: Panels-in-Series", "OPT_MATH: Array Area (m²)",
    "OPT_MATH: Budget Binding?", "OPT_MATH: Solver Status",
]
REC_MATH_HEADERS = [
    "REC_MATH: Panels", "REC_MATH: System kW DC", "REC_MATH: Annual Prod (kWh)",
    "REC_MATH: Gross CAPEX ($)", "REC_MATH: Net CAPEX ITC ($)",
    "REC_MATH: Annual Savings ($)", "REC_MATH: Payback (yr)", "REC_MATH: NPV ($)",
    "REC_MATH: String Voltage (V)", "REC_MATH: Parallel Strings",
    "REC_MATH: Panels-in-Series", "REC_MATH: Array Area (m²)",
    "REC_MATH: Budget Binding?", "REC_MATH: Solver Status",
    "REC_MATH: Battery kWh", "REC_MATH: Battery Model",
]
LLM_OPT_HEADERS = [
    "LLM_OPT: Panels", "LLM_OPT: System kW DC", "LLM_OPT: Target Offset (%)",
    "LLM_OPT: Annual Prod (kWh)", "LLM_OPT: Annual Savings ($)",
    "LLM_OPT: CAPEX ($)", "LLM_OPT: Payback (yr)", "LLM_OPT: Budget Binding?",
]
LLM_REC_HEADERS = [
    "LLM_REC: Panels", "LLM_REC: System kW DC", "LLM_REC: Target Offset (%)",
    "LLM_REC: Annual Prod (kWh)", "LLM_REC: Annual Savings ($)",
    "LLM_REC: CAPEX ($)", "LLM_REC: Payback (yr)", "LLM_REC: Budget Binding?",
]
ALL_HEADERS = (INPUT_HEADERS + MATH_HEADERS + OPT_MATH_HEADERS +
               REC_MATH_HEADERS + LLM_OPT_HEADERS + LLM_REC_HEADERS)

SUMMARY_HEADERS = [
    "File #", "Location", "Budget (USD)", "Roof Area (m²)",
    "Roof Length (m)", "Roof Width (m)", "Panel Brand",
    "Num EVs", "Num People", "Num Daytime Occupants",
    "MATH_OPT: Gross CAPEX ($)",  "MATH_OPT: Annual Savings ($)",
    "MATH_REC: Gross CAPEX ($)",  "MATH_REC: Annual Savings ($)",
    "LLM_OPT: CAPEX ($)",         "LLM_OPT: Annual Savings ($)",
    "LLM_REC: CAPEX ($)",         "LLM_REC: Annual Savings ($)",
]


# ══════════════════════════════════════════════════════════════════════════════
# STYLING
# ══════════════════════════════════════════════════════════════════════════════

C_INPUT="D9E1F2"; C_MATH="E2EFDA"; C_OPT_M="FFF2CC"
C_REC_M="FCE4D6"; C_LLM_O="DDEBF7"; C_LLM_R="F4CCFF"
C_TITLE="1F4E79"; C_WHITE="FFFFFF"

COL_COLORS = ([C_INPUT]*len(INPUT_HEADERS) + [C_MATH]*len(MATH_HEADERS) +
              [C_OPT_M]*len(OPT_MATH_HEADERS) + [C_REC_M]*len(REC_MATH_HEADERS) +
              [C_LLM_O]*len(LLM_OPT_HEADERS)  + [C_LLM_R]*len(LLM_REC_HEADERS))

GROUPS = [
    ("INPUTS",                          len(INPUT_HEADERS),    C_INPUT),
    ("PYTHON — PANEL SPEC",             len(MATH_HEADERS),     C_MATH),
    ("PYTHON — OPTIMAL (MILP)",         len(OPT_MATH_HEADERS), C_OPT_M),
    ("PYTHON — RECOMMENDED (MILP+BAT)", len(REC_MATH_HEADERS), C_REC_M),
    ("LLM — OPTIMAL SYSTEM",            len(LLM_OPT_HEADERS),  C_LLM_O),
    ("LLM — RECOMMENDED SYSTEM",        len(LLM_REC_HEADERS),  C_LLM_R),
]

SUM_COLORS = ([C_INPUT]*10 + [C_OPT_M,C_OPT_M] + [C_REC_M,C_REC_M] +
              [C_LLM_O,C_LLM_O] + [C_LLM_R,C_LLM_R])
SUM_GROUPS = [
    ("INPUTS",               10, C_INPUT),
    ("PYTHON — OPTIMAL",      2, C_OPT_M),
    ("PYTHON — RECOMMENDED",  2, C_REC_M),
    ("LLM — OPTIMAL",         2, C_LLM_O),
    ("LLM — RECOMMENDED",     2, C_LLM_R),
]

def _hf(c): return PatternFill("solid", fgColor=c)


def _apply_header(ws, groups, col_colors, all_hdrs):
    col = 1
    for label, span, color in groups:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + span - 1)
        c = ws.cell(row=1, column=col, value=label)
        c.font      = Font(bold=True, color=C_WHITE, name="Arial", size=10)
        c.fill      = _hf(C_TITLE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        col += span
    for ci, (hdr, color) in enumerate(zip(all_hdrs, col_colors), start=1):
        c = ws.cell(row=2, column=ci, value=hdr)
        c.font      = Font(bold=True, name="Arial", size=9)
        c.fill      = _hf(color)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 42


def _fmt(cell, hdr):
    if any(x in hdr for x in ["($)", "CAPEX", "Savings", "NPV"]):
        cell.number_format = '$#,##0.00'
    elif "(kWh)" in hdr or "kWh" in hdr:
        cell.number_format = '#,##0.0'
    elif "kW DC" in hdr:
        cell.number_format = '0.000'
    elif "(m²)" in hdr:
        cell.number_format = '0.00'
    elif "(%)" in hdr or "Efficiency" in hdr:
        cell.number_format = '0.0'
    elif "Payback" in hdr:
        cell.number_format = '0.0'


# ══════════════════════════════════════════════════════════════════════════════
# SHEET WRITER
# ══════════════════════════════════════════════════════════════════════════════

def write_sheet(ws, manifest_df, output_dir, sheet_label):
    ws.freeze_panes = "B3"
    _apply_header(ws, GROUPS, COL_COLORS, ALL_HEADERS)

    for ci, hdr in enumerate(ALL_HEADERS, start=1):
        w = (30 if any(x in hdr for x in ["Selected Panel", "Battery Model", "Status"])
             else 16 if any(x in hdr for x in ["Location", "Rate Plan"])
             else 20)
        ws.column_dimensions[get_column_letter(ci)].width = w

    seen = set(); rows_written = 0; summary_rows = []

    for _, mrow in manifest_df.iterrows():
        fn = int(mrow["file_num"])
        if fn in seen: continue
        seen.add(fn); rows_written += 1
        excel_row = rows_written + 2

        t0 = time.time()
        print(f"  [{sheet_label}] file {fn:3d} ({rows_written}/"
              f"{len(manifest_df.drop_duplicates('file_num'))}) — MILP ...",
              end=" ", flush=True)
        try:
            math_res = run_optimizer_for_row(mrow.to_dict())
            print(f"done in {time.time()-t0:.1f}s | "
                  f"OPT={math_res['opt_panels']}p REC={math_res['rec_panels']}p "
                  f"bat={math_res['rec_bat_model']}")
        except Exception as e:
            import traceback; traceback.print_exc()
            print(f"ERROR: {e}")
            math_res = None

        txt_path = os.path.join(output_dir, f"{fn}.txt")
        parsed   = parse_txt(txt_path)
        o = parsed["optimal"]     if parsed else {}
        r = parsed["recommended"] if parsed else {}
        gv = lambda d, k: d.get(k)

        if math_res:
            values = [
                fn, mrow["location"], mrow["latitude"], mrow["longitude"],
                mrow["budget_usd"], mrow["roof_area_m2"],
                mrow["roof_length_m"], mrow["roof_breadth_m"],
                mrow["rate_plan"], mrow["panel_brand"],
                mrow["num_evs"], mrow["num_people"], mrow["num_daytime_occupants"],
                math_res["panel_name"], math_res["panel_power_w"], math_res["panel_eff_pct"],
                math_res["panel_area_m2"], math_res["cell_vmp_v"], math_res["cell_imp_a"],
                math_res["max_panels_roof"], math_res["max_panels_budget"],
                math_res["prod_per_panel"], math_res["annual_kwh_auto"],
                math_res["irradiance"], math_res["annual_load_kwh"],
                math_res["opt_panels"], math_res["opt_kw_dc"], math_res["opt_prod_kwh"],
                math_res["opt_gross_capex"], math_res["opt_net_capex"],
                math_res["opt_savings"], math_res["opt_payback"], math_res["opt_npv"],
                math_res["opt_sv"], math_res["opt_snp"], math_res["opt_sns"],
                math_res["opt_array_area"], math_res["opt_budget_bind"], math_res["opt_status"],
                math_res["rec_panels"], math_res["rec_kw_dc"], math_res["rec_prod_kwh"],
                math_res["rec_gross_capex"], math_res["rec_net_capex"],
                math_res["rec_savings"], math_res["rec_payback"], math_res["rec_npv"],
                math_res["rec_sv"], math_res["rec_snp"], math_res["rec_sns"],
                math_res["rec_array_area"], math_res["rec_budget_bind"], math_res["rec_status"],
                math_res["rec_bat_kwh"], math_res["rec_bat_model"],
                gv(o, "panels"), gv(o, "kw_dc"), gv(o, "offset_pct"),
                gv(o, "annual_prod_kwh"), gv(o, "annual_savings_usd"),
                gv(o, "capex_usd"), gv(o, "payback_yr"), gv(o, "budget_binding"),
                gv(r, "panels"), gv(r, "kw_dc"), gv(r, "offset_pct"),
                gv(r, "annual_prod_kwh"), gv(r, "annual_savings_usd"),
                gv(r, "capex_usd"), gv(r, "payback_yr"), gv(r, "budget_binding"),
            ]
            summary_rows.append([
                fn, mrow["location"], mrow["budget_usd"], mrow["roof_area_m2"],
                mrow["roof_length_m"], mrow["roof_breadth_m"], mrow["panel_brand"],
                mrow["num_evs"], mrow["num_people"], mrow["num_daytime_occupants"],
                math_res["opt_gross_capex"], math_res["opt_savings"],
                math_res["rec_gross_capex"], math_res["rec_savings"],
                gv(o, "capex_usd"), gv(o, "annual_savings_usd"),
                gv(r, "capex_usd"), gv(r, "annual_savings_usd"),
            ])
        else:
            inp = list(mrow[["file_num", "location", "latitude", "longitude",
                              "budget_usd", "roof_area_m2", "roof_length_m", "roof_breadth_m",
                              "rate_plan", "panel_brand", "num_evs", "num_people",
                              "num_daytime_occupants"]])
            values = inp + [None] * (len(ALL_HEADERS) - 13)
            summary_rows.append(
                [fn, mrow["location"], mrow["budget_usd"], mrow["roof_area_m2"],
                 mrow["roof_length_m"], mrow["roof_breadth_m"], mrow["panel_brand"],
                 mrow["num_evs"], mrow["num_people"], mrow["num_daytime_occupants"]]
                + [None] * 8)

        for ci, (val, color) in enumerate(zip(values, COL_COLORS), start=1):
            c = ws.cell(row=excel_row, column=ci, value=val)
            c.font      = Font(name="Arial", size=9)
            c.fill      = _hf(color)
            c.alignment = Alignment(horizontal="center", vertical="center")
            _fmt(c, ALL_HEADERS[ci - 1])

    print(f"  [{sheet_label}] Complete — {rows_written} unique rows.\n")
    return summary_rows


# ══════════════════════════════════════════════════════════════════════════════
# SUMMARY WRITER
# ══════════════════════════════════════════════════════════════════════════════

def write_summary_sheet(ws, rows, sheet_label):
    ws.freeze_panes = "B3"
    _apply_header(ws, SUM_GROUPS, SUM_COLORS, SUMMARY_HEADERS)
    for ci in range(1, len(SUMMARY_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 24
    for ri, row in enumerate(rows, start=3):
        for ci, (val, color) in enumerate(zip(row, SUM_COLORS), start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font      = Font(name="Arial", size=9)
            c.fill      = _hf(color)
            c.alignment = Alignment(horizontal="center", vertical="center")
            hdr = SUMMARY_HEADERS[ci - 1]
            if any(x in hdr for x in ["($)", "CAPEX", "Savings"]):
                c.number_format = '$#,##0.00'
    print(f"  Summary sheet '{sheet_label}' written — {len(rows)} rows.")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main():
    print("=" * 65)
    print("  Baseline Comparison Builder  (full MILP per row)")
    print("=" * 65)
    print(f"  Battery options in MILP: {len(BATTERY_OPTIONS)} "
          f"(none + {len(BATTERY_CATALOG)} models × {MAX_BATTERY_UNITS} unit counts)")
    print(f"  Per-option eta used in SOC dynamics (disaggregated).")
    print()

    df1 = pd.read_csv(MANIFEST_OP1)
    df2 = pd.read_csv(MANIFEST_OP2).drop_duplicates(subset=["file_num"], keep="first")
    print(f"op1: {len(df1)} rows   op2: {len(df2)} unique rows\n")

    wb  = Workbook()
    ws1 = wb.active; ws1.title = "op1 (72 rows)"
    ws2 = wb.create_sheet("op2 (243 rows)")

    print("▶  Processing op1 ...")
    sum1 = write_sheet(ws1, df1, OUTPUT_DIR_1, "op1")

    print("▶  Processing op2 ...")
    sum2 = write_sheet(ws2, df2, OUTPUT_DIR_2, "op2")

    os.makedirs(os.path.dirname(OUTPUT_XLSX), exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(f"\n✓ Detail file saved:\n  {OUTPUT_XLSX}")

    wb_s = Workbook()
    wb_s.remove(wb_s.active)
    ws_s1 = wb_s.create_sheet("Summary op1")
    ws_s2 = wb_s.create_sheet("Summary op2")
    write_summary_sheet(ws_s1, sum1, "op1")
    write_summary_sheet(ws_s2, sum2, "op2")
    wb_s.save(SUMMARY_XLSX)
    print(f"✓ Summary file saved:\n  {SUMMARY_XLSX}")


if __name__ == "__main__":
    main()