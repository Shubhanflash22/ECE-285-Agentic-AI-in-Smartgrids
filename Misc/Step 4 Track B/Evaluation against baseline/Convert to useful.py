"""
generate_total_comparison.py

Reads both sheets from baseline_comparison.xlsx and produces a Total sheet
matching the format of Total_Comparison_old.xlsx (Total tab).

- Columns A-I: input fields pulled directly from baseline data
- Columns J, K: PYTHON OPTIMAL Gross CAPEX and Annual Savings (hard values)
- Columns L, M: LLM OPTIMAL CAPEX and Annual Savings (hard values)
- Column N: % change formula =(L-J)/J  (Optimal CAPEX % diff)
- Column O: % change formula =(M-K)/K  (Optimal Savings % diff)
- Columns P, Q: PYTHON RECOMMENDED Gross CAPEX and Annual Savings (hard values)
- Columns R, S: LLM RECOMMENDED CAPEX and Annual Savings (hard values)
- Column T: % change formula =(R-P)/P  (Recommended CAPEX % diff)
- Column U: % change formula =(S-Q)/Q  (Recommended Savings % diff)

Usage:
    python generate_total_comparison.py baseline_comparison.xlsx output.xlsx
"""

import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Column indices in baseline sheets (1-based) ──────────────────────────────
SRC = {
    "Location":                2,   # B
    "Budget (USD)":            5,   # E
    "Roof Area (m²)":          6,   # F
    "Roof Length (m)":         7,   # G
    "Roof Width (m)":          8,   # H
    "Panel Brand":            10,   # J
    "Num EVs":                11,   # K
    "Num People":             12,   # L
    "Num Daytime Occupants":  13,   # M
    "OPT_MATH: Gross CAPEX":  29,   # AC
    "OPT_MATH: Annual Savings": 31, # AE
    "LLM_OPT: CAPEX":         61,   # BI
    "LLM_OPT: Annual Savings": 60,  # BH
    "REC_MATH: Gross CAPEX":  43,   # AQ
    "REC_MATH: Annual Savings": 45, # AS
    "LLM_REC: CAPEX":         69,   # BQ
    "LLM_REC: Annual Savings": 68,  # BP
}

# ── Output column layout (1-based) ───────────────────────────────────────────
# A=1 … U=21
OUT_COLS = {
    "Location": 1, "Budget (USD)": 2, "Roof Area (m²)": 3,
    "Roof Length (m)": 4, "Roof Width (m)": 5, "Panel Brand": 6,
    "Num EVs": 7, "Num People": 8, "Num Daytime Occupants": 9,
    "MATH_OPT: Gross CAPEX ($)": 10,
    "MATH_OPT: Annual Savings ($)": 11,
    "LLM_OPT: CAPEX ($)": 12,
    "LLM_OPT: Annual Savings ($)": 13,
    "Gross CAPEX ($)": 14,       # formula col N
    "Annual Savings ($)_opt": 15, # formula col O
    "MATH_REC: Gross CAPEX ($)": 16,
    "MATH_REC: Annual Savings ($)": 17,
    "LLM_REC: CAPEX ($)": 18,
    "LLM_REC: Annual Savings ($)": 19,
    "CAPEX ($)": 20,             # formula col T
    "Annual Savings ($)_rec": 21, # formula col U
}

INPUT_KEYS = [
    "Location", "Budget (USD)", "Roof Area (m²)", "Roof Length (m)",
    "Roof Width (m)", "Panel Brand", "Num EVs", "Num People",
    "Num Daytime Occupants"
]

# ── Styling helpers ───────────────────────────────────────────────────────────
HEADER1_FILL = PatternFill("solid", start_color="4472C4", end_color="4472C4")
HEADER2_FILL = PatternFill("solid", start_color="8EA9DB", end_color="8EA9DB")
FORMULA_FILL = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
WHITE_FONT   = Font(name="Arial", bold=True, color="FFFFFF")
BOLD_FONT    = Font(name="Arial", bold=True)
REG_FONT     = Font(name="Arial")
CENTER       = Alignment(horizontal="center", vertical="center")
PCT_FMT      = "0.0%"
NUM_FMT      = '#,##0.00'

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def apply_header1(cell, value, span_hint=None):
    cell.value = value
    cell.font = WHITE_FONT
    cell.fill = HEADER1_FILL
    cell.alignment = CENTER

def apply_header2(cell, value):
    cell.value = value
    cell.font = BOLD_FONT
    cell.fill = HEADER2_FILL
    cell.alignment = CENTER

def build_total_sheet(ws_out, all_rows):
    """Write headers + data rows with Excel formulas into ws_out."""

    # ── Row 1: section headers ────────────────────────────────────────────────
    sections = [
        (1,  9,  "INPUTS"),
        (10, 11, "PYTHON — OPTIMAL"),
        (12, 13, "LLM — OPTIMAL"),
        (14, 15, "Optimal"),
        (16, 17, "PYTHON — RECOMMENDED"),
        (18, 19, "LLM — RECOMMENDED"),
        (20, 21, "Recommended"),
    ]
    for start_c, end_c, label in sections:
        cell = ws_out.cell(row=1, column=start_c)
        apply_header1(cell, label)
        if end_c > start_c:
            ws_out.merge_cells(
                start_row=1, start_column=start_c,
                end_row=1,   end_column=end_c
            )

    # ── Row 2: column headers ─────────────────────────────────────────────────
    col_labels = [
        "Location", "Budget (USD)", "Roof Area (m²)", "Roof Length (m)",
        "Roof Width (m)", "Panel Brand", "Num EVs", "Num People",
        "Num Daytime Occupants",
        "MATH_OPT: Gross CAPEX ($)", "MATH_OPT: Annual Savings ($)",
        "LLM_OPT: CAPEX ($)",        "LLM_OPT: Annual Savings ($)",
        "Gross CAPEX ($)",           "Annual Savings ($)",
        "MATH_REC: Gross CAPEX ($)", "MATH_REC: Annual Savings ($)",
        "LLM_REC: CAPEX ($)",        "LLM_REC: Annual Savings ($)",
        "CAPEX ($)",                 "Annual Savings ($)",
    ]
    for c_idx, label in enumerate(col_labels, start=1):
        cell = ws_out.cell(row=2, column=c_idx)
        apply_header2(cell, label)

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_offset, row_data in enumerate(all_rows):
        r = row_offset + 3  # data starts at row 3

        # Input fields A-I
        for key in INPUT_KEYS:
            c = OUT_COLS[key]
            ws_out.cell(row=r, column=c).value = row_data[key]
            ws_out.cell(row=r, column=c).font = REG_FONT

        # Hard values J-M, P-S
        value_map = [
            (10, "OPT_MATH: Gross CAPEX"),
            (11, "OPT_MATH: Annual Savings"),
            (12, "LLM_OPT: CAPEX"),
            (13, "LLM_OPT: Annual Savings"),
            (16, "REC_MATH: Gross CAPEX"),
            (17, "REC_MATH: Annual Savings"),
            (18, "LLM_REC: CAPEX"),
            (19, "LLM_REC: Annual Savings"),
        ]
        for c, key in value_map:
            cell = ws_out.cell(row=r, column=c)
            cell.value = row_data[key]
            cell.number_format = NUM_FMT
            cell.font = REG_FONT

        # Formula columns N, O, T, U
        J = get_column_letter(10)
        K = get_column_letter(11)
        L = get_column_letter(12)
        M = get_column_letter(13)
        P = get_column_letter(16)
        Q = get_column_letter(17)
        Rc = get_column_letter(18)
        S = get_column_letter(19)

        formulas = {
            14: f"=({L}{r} - {J}{r}) / {J}{r}",   # Optimal CAPEX % diff
            15: f"=({M}{r} - {K}{r}) / {K}{r}",   # Optimal Savings % diff
            20: f"=({Rc}{r} - {P}{r}) / {P}{r}",  # Recommended CAPEX % diff
            21: f"=({S}{r} - {Q}{r}) / {Q}{r}",   # Recommended Savings % diff
        }
        for c, formula in formulas.items():
            cell = ws_out.cell(row=r, column=c)
            cell.value = formula
            cell.number_format = PCT_FMT
            cell.fill = FORMULA_FILL
            cell.font = REG_FONT

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = [15, 12, 13, 14, 13, 16, 9, 11, 20,
              22, 23, 17, 22, 14, 16,
              24, 25, 17, 23, 12, 16]
    for i, w in enumerate(widths, start=1):
        ws_out.column_dimensions[get_column_letter(i)].width = w

    # Freeze panes below headers
    ws_out.freeze_panes = "A3"


def read_baseline_rows(wb_in):
    """Combine rows from both sheets; skip header rows (rows 1-2)."""
    rows = []
    for sname in wb_in.sheetnames:
        ws = wb_in[sname]
        for r in range(3, ws.max_row + 1):
            row = {
                key: ws.cell(row=r, column=col).value
                for key, col in SRC.items()
            }
            # Skip completely empty rows
            if all(v is None for v in row.values()):
                continue
            rows.append(row)
    return rows


def main(input_path, output_path):
    wb_in = openpyxl.load_workbook(input_path)
    all_rows = read_baseline_rows(wb_in)
    print(f"Read {len(all_rows)} data rows from {input_path}")

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "Total"

    build_total_sheet(ws_out, all_rows)

    wb_out.save(output_path)
    print(f"Saved output to {output_path}")
    print(f"Total data rows written: {len(all_rows)}")


if __name__ == "__main__":
    ip = r"C:\Users\shubh\Downloads\baseline_comparison.xlsx"
    op = r"C:\Users\shubh\Downloads\Total Comparison generated.xlsx"
    main(ip, op)